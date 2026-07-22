from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber


ROOT = Path(__file__).resolve().parents[1]
REFERENCES = ROOT / "referencias"
OUTPUT = ROOT / "tmp" / "reference-inspection"
PDF_PATH = REFERENCES / "Prompt_Formulas_Conceptos.pdf"
XLSX_PATH = REFERENCES / "Formulas.Comercio.xlsx"

REFERENCE_RE = re.compile(r"\b([NIARUL])\s*\[\s*([^\]]*)\s*\]")
FORMULA_TOKEN_RE = re.compile(r"[NIARUL]\s*\[|SI\s*\(|\bY\s*\(|\bO\s*\(|\bNO\s*\(|&&|\|\||[+\-*/<>!=]=?", re.IGNORECASE)


def cell_to_value(cell: Any) -> Any:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def row_values(row: tuple[Any, ...]) -> list[Any]:
    return [cell_to_value(cell) for cell in row]


def first_non_empty_rows(sheet: Any, limit: int = 12) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows():
        values = row_values(row)
        if any(value is not None for value in values):
            rows.append(
                {
                    "row": row[0].row,
                    "values": values[: min(len(values), 40)],
                }
            )
        if len(rows) >= limit:
            break
    return rows


def likely_header_row(sheet: Any) -> int | None:
    best_row: int | None = None
    best_score = 0
    max_row = min(sheet.max_row or 1, 25)
    for row in sheet.iter_rows(min_row=1, max_row=max_row):
        values = row_values(row)
        non_empty = [value for value in values if value is not None]
        textish = [value for value in non_empty if isinstance(value, str)]
        score = len(non_empty) + len(textish)
        if score > best_score:
            best_score = score
            best_row = row[0].row
    return best_row


def sheet_profile(sheet: Any) -> dict[str, Any]:
    dimensions = sheet.calculate_dimension(force=True)
    header_row = likely_header_row(sheet)
    headers: list[Any] = []
    duplicate_headers: list[str] = []
    if header_row:
        headers = row_values(next(sheet.iter_rows(min_row=header_row, max_row=header_row)))
        seen: Counter[str] = Counter(
            str(value).strip().casefold() for value in headers if value is not None
        )
        duplicate_headers = [header for header, count in seen.items() if count > 1]

    non_empty_rows = 0
    formula_like_cells: list[dict[str, Any]] = []
    references_by_type: dict[str, Counter[str]] = defaultdict(Counter)
    column_non_empty: Counter[str] = Counter()
    sample_rows_after_header: list[dict[str, Any]] = []

    for row in sheet.iter_rows():
        values = row_values(row)
        if any(value is not None for value in values):
            non_empty_rows += 1
            if header_row and row[0].row > header_row and len(sample_rows_after_header) < 8:
                sample_rows_after_header.append(
                    {
                        "row": row[0].row,
                        "values": values[: min(len(values), 40)],
                    }
                )
        for cell in row:
            value = cell_to_value(cell)
            if value is None:
                continue
            column_non_empty[cell.column_letter] += 1
            if isinstance(value, str) and FORMULA_TOKEN_RE.search(value):
                refs = REFERENCE_RE.findall(value)
                for ref_type, ref_id in refs:
                    references_by_type[ref_type.upper()][ref_id.strip()] += 1
                if len(formula_like_cells) < 40:
                    formula_like_cells.append(
                        {
                            "cell": cell.coordinate,
                            "row": cell.row,
                            "column": cell.column_letter,
                            "value": value,
                            "references": [
                                {"type": ref_type.upper(), "id": ref_id.strip()}
                                for ref_type, ref_id in refs
                            ],
                        }
                    )

    return {
        "name": sheet.title,
        "dimensions": dimensions,
        "maxRow": sheet.max_row,
        "maxColumn": sheet.max_column,
        "nonEmptyRows": non_empty_rows,
        "likelyHeaderRow": header_row,
        "headers": headers,
        "duplicateHeaders": duplicate_headers,
        "firstNonEmptyRows": first_non_empty_rows(sheet),
        "sampleRowsAfterHeader": sample_rows_after_header,
        "columnNonEmptyCounts": dict(column_non_empty),
        "formulaLikeCellCountSampled": len(formula_like_cells),
        "formulaLikeCells": formula_like_cells,
        "referencesByType": {
            ref_type: refs.most_common(40) for ref_type, refs in references_by_type.items()
        },
    }


def extract_pdf() -> dict[str, Any]:
    pages: list[dict[str, Any]] = []
    all_text: list[str] = []
    with pdfplumber.open(PDF_PATH) as pdf:
        for index, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            all_text.append(f"\n\n--- PAGINA {index} ---\n{text}")
            pages.append(
                {
                    "page": index,
                    "width": page.width,
                    "height": page.height,
                    "textLength": len(text),
                    "preview": text[:2000],
                }
            )
    text = "".join(all_text)
    (OUTPUT / "pdf_text.txt").write_text(text, encoding="utf-8")
    headings = []
    for line in text.splitlines():
        clean = line.strip()
        if clean and (clean.isupper() or clean.startswith(("Hoja", "Columnas", "Fórmulas", "Formulas"))):
            headings.append(clean)
    return {
        "path": str(PDF_PATH),
        "pageCount": len(pages),
        "pages": pages,
        "headingsSample": headings[:80],
        "referenceMentions": Counter(match.group(0) for match in REFERENCE_RE.finditer(text)).most_common(80),
    }


def inspect_workbook() -> dict[str, Any]:
    workbook = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=False)
    profiles = [sheet_profile(workbook[sheet_name]) for sheet_name in workbook.sheetnames]
    workbook.close()
    return {
        "path": str(XLSX_PATH),
        "sheetNames": [profile["name"] for profile in profiles],
        "sheets": profiles,
    }


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    result = {
        "pdf": extract_pdf(),
        "workbook": inspect_workbook(),
    }
    (OUTPUT / "inspection.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(json.dumps({
        "pdfPages": result["pdf"]["pageCount"],
        "sheetNames": result["workbook"]["sheetNames"],
        "output": str(OUTPUT),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
