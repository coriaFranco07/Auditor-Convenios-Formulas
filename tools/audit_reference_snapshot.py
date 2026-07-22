from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "referencias" / "Formulas.Comercio.xlsx"
OUTPUT = ROOT / "tmp" / "reference-inspection" / "reference_issues.json"

REFERENCE_RE = re.compile(r"\b([NIARUL])\s*\[\s*([^\]]*)\s*\]", re.IGNORECASE)
FORMULA_CHARS_RE = re.compile(r"[NIARUL]\s*\[|SI\s*\(|\bY\s*\(|\bO\s*\(|\bNO\s*\(|&&|\|\||[+\-*/<>!=]=?", re.IGNORECASE)
BOOLEAN_HINT_RE = re.compile(r"(>=|<=|==|!=|>|<|&&|\|\||\bSI\s*\(|\bY\s*\(|\bO\s*\(|\bNO\s*\(|!)", re.IGNORECASE)


SCHEMA = {
    "concepts": {
        "sheet": "Conceptos y Formulas (1)",
        "header_row": 2,
        "columns": {
            "id": "A",
            "name": "B",
            "activation": "C",
            "scope": "D",
            "monthly_condition": "E",
            "monthly_true": "F",
            "monthly_false": "G",
            "monthly_unit": "H",
            "daily_condition": "I",
            "daily_true": "J",
            "daily_false": "K",
            "daily_unit": "L",
            "totalizes": "M",
            "pre_formula": "N",
            "post_formula": "O",
        },
        "formula_columns": ["E", "F", "G", "H", "I", "J", "K", "L", "N", "O"],
        "condition_columns": ["E", "I"],
    },
    "variables": {
        "sheet": "Variables de Legajos (2)",
        "header_row": 2,
        "columns": {"id": "A", "name": "B", "abbr": "C"},
    },
    "auxiliaries": {
        "sheet": "Calculo Auxiliares (3)",
        "header_row": 1,
        "columns": {
            "id": "A",
            "name": "B",
            "true_formula": "C",
            "false_formula": "D",
            "condition": "E",
            "value": "F",
            "class": "G",
        },
        "formula_columns": ["C", "D", "E"],
        "condition_columns": ["E"],
    },
    "accumulators": {
        "sheet": "Acumuladores (4)",
        "header_row": 1,
        "columns": {"id": "A", "name": "B", "concept_id": "C", "concept_name": "D", "operation": "E"},
    },
    "conventions": {
        "sheet": "Convenios (5)",
        "header_row": 1,
        "columns": {"id": "A", "name": "B"},
    },
}


def value(cell: Any) -> Any:
    raw = cell.value
    if isinstance(raw, str):
        clean = raw.strip()
        return clean if clean else None
    return raw


def as_int(raw: Any) -> int | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float) and raw.is_integer():
        return int(raw)
    if isinstance(raw, str):
        clean = raw.strip()
        if re.fullmatch(r"\d+", clean):
            return int(clean)
    return None


def row_is_empty(sheet: Any, row_number: int) -> bool:
    return all(value(cell) is None for cell in sheet[row_number])


def collect_records(sheet: Any, header_row: int, columns: dict[str, str]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        if row_is_empty(sheet, row_number):
            continue
        record = {"row": row_number, "raw": {}}
        for key, column in columns.items():
            record[key] = value(sheet[f"{column}{row_number}"])
            record["raw"][key] = record[key]
        records.append(record)
    return records


def formula_cells(sheet: Any, start_row: int, columns: Iterable[str]) -> list[dict[str, Any]]:
    cells = []
    for row_number in range(start_row, sheet.max_row + 1):
        if row_is_empty(sheet, row_number):
            continue
        for column in columns:
            cell = sheet[f"{column}{row_number}"]
            raw = value(cell)
            if isinstance(raw, str) and FORMULA_CHARS_RE.search(raw):
                cells.append(
                    {
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "row": row_number,
                        "column": column,
                        "formula": raw,
                        "references": [
                            {"type": ref_type.upper(), "id": ref_id.strip()}
                            for ref_type, ref_id in REFERENCE_RE.findall(raw)
                        ],
                    }
                )
    return cells


def duplicate_groups(records: list[dict[str, Any]], id_key: str) -> list[dict[str, Any]]:
    groups: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        record_id = as_int(record.get(id_key))
        if record_id is not None:
            groups[record_id].append(record)
    result = []
    for record_id, definitions in groups.items():
        if len(definitions) < 2:
            continue
        signatures = Counter(json.dumps(definition["raw"], ensure_ascii=False, sort_keys=True, default=str) for definition in definitions)
        result.append(
            {
                "id": record_id,
                "count": len(definitions),
                "kind": "identical" if len(signatures) == 1 else "conflict",
                "rows": [definition["row"] for definition in definitions],
            }
        )
    return result


def build_dependency_graph(cells: list[dict[str, Any]], concept_rows: dict[int, int], auxiliary_rows: dict[int, int]) -> dict[str, set[str]]:
    graph: dict[str, set[str]] = defaultdict(set)
    for cell in cells:
        owner: str | None = None
        if cell["sheet"] == SCHEMA["concepts"]["sheet"]:
            for concept_id, row in concept_rows.items():
                if row == cell["row"]:
                    owner = f"R[{concept_id}]"
                    break
        elif cell["sheet"] == SCHEMA["auxiliaries"]["sheet"]:
            for auxiliary_id, row in auxiliary_rows.items():
                if row == cell["row"]:
                    owner = f"A[{auxiliary_id}]"
                    break
        if not owner:
            continue
        graph[owner]
        for ref in cell["references"]:
            if ref["type"] in {"A", "R"} and re.fullmatch(r"\d+", ref["id"]):
                graph[owner].add(f"{ref['type']}[{int(ref['id'])}]")
    return graph


def strongly_connected(graph: dict[str, set[str]]) -> list[list[str]]:
    index = 0
    stack: list[str] = []
    indices: dict[str, int] = {}
    lowlinks: dict[str, int] = {}
    on_stack: set[str] = set()
    components: list[list[str]] = []

    def visit(node: str) -> None:
        nonlocal index
        indices[node] = index
        lowlinks[node] = index
        index += 1
        stack.append(node)
        on_stack.add(node)
        for target in graph.get(node, set()):
            if target not in indices:
                visit(target)
                lowlinks[node] = min(lowlinks[node], lowlinks[target])
            elif target in on_stack:
                lowlinks[node] = min(lowlinks[node], indices[target])
        if lowlinks[node] == indices[node]:
            component = []
            while True:
                target = stack.pop()
                on_stack.remove(target)
                component.append(target)
                if target == node:
                    break
            components.append(component)

    all_nodes = set(graph)
    for targets in graph.values():
        all_nodes.update(targets)
    for node in sorted(all_nodes):
        if node not in indices:
            visit(node)
    return [
        component
        for component in components
        if len(component) > 1 or (len(component) == 1 and component[0] in graph.get(component[0], set()))
    ]


def main() -> None:
    workbook = openpyxl.load_workbook(XLSX_PATH, read_only=False, data_only=False)

    concept_sheet = workbook[SCHEMA["concepts"]["sheet"]]
    variable_sheet = workbook[SCHEMA["variables"]["sheet"]]
    auxiliary_sheet = workbook[SCHEMA["auxiliaries"]["sheet"]]
    accumulator_sheet = workbook[SCHEMA["accumulators"]["sheet"]]
    convention_sheet = workbook[SCHEMA["conventions"]["sheet"]]

    concepts = collect_records(concept_sheet, SCHEMA["concepts"]["header_row"], SCHEMA["concepts"]["columns"])
    variables = collect_records(variable_sheet, SCHEMA["variables"]["header_row"], SCHEMA["variables"]["columns"])
    auxiliaries = collect_records(auxiliary_sheet, SCHEMA["auxiliaries"]["header_row"], SCHEMA["auxiliaries"]["columns"])
    accumulators = collect_records(accumulator_sheet, SCHEMA["accumulators"]["header_row"], SCHEMA["accumulators"]["columns"])
    conventions = collect_records(convention_sheet, SCHEMA["conventions"]["header_row"], SCHEMA["conventions"]["columns"])

    concept_ids = {as_int(record["id"]) for record in concepts if as_int(record["id"]) is not None}
    variable_ids = {as_int(record["id"]) for record in variables if as_int(record["id"]) is not None}
    auxiliary_ids = {as_int(record["id"]) for record in auxiliaries if as_int(record["id"]) is not None}
    concept_rows = {as_int(record["id"]): record["row"] for record in concepts if as_int(record["id"]) is not None}
    auxiliary_rows = {as_int(record["id"]): record["row"] for record in auxiliaries if as_int(record["id"]) is not None}

    cells = formula_cells(concept_sheet, 3, SCHEMA["concepts"]["formula_columns"])
    cells += formula_cells(auxiliary_sheet, 2, SCHEMA["auxiliaries"]["formula_columns"])

    missing = []
    malformed_refs = []
    for cell in cells:
        for ref in cell["references"]:
            if not re.fullmatch(r"\d+", ref["id"]):
                malformed_refs.append({**cell, "reference": ref})
                continue
            ref_id = int(ref["id"])
            expected = {
                "L": variable_ids,
                "A": auxiliary_ids,
                "R": concept_ids,
                "U": concept_ids,
                "N": concept_ids,
                "I": concept_ids,
            }[ref["type"]]
            if ref_id not in expected:
                missing.append({**cell, "reference": {"type": ref["type"], "id": ref_id}})

    numeric_conditions = []
    condition_columns = {
        (SCHEMA["concepts"]["sheet"], "E"),
        (SCHEMA["concepts"]["sheet"], "I"),
        (SCHEMA["auxiliaries"]["sheet"], "E"),
    }
    for cell in cells:
        if (cell["sheet"], cell["column"]) in condition_columns and not BOOLEAN_HINT_RE.search(cell["formula"]):
            numeric_conditions.append(cell)

    graph = build_dependency_graph(cells, concept_rows, auxiliary_rows)
    cycles = strongly_connected(graph)

    result = {
        "counts": {
            "concepts": len(concepts),
            "variables": len(variables),
            "auxiliaries": len(auxiliaries),
            "accumulators": len(accumulators),
            "conventions": len(conventions),
            "formulaCells": len(cells),
            "missingReferences": len(missing),
            "malformedReferences": len(malformed_refs),
            "numericConditions": len(numeric_conditions),
            "cycles": len(cycles),
        },
        "catalogValues": {
            "activation": sorted({str(record["activation"]) for record in concepts if record.get("activation") is not None}),
            "scope": sorted({str(record["scope"]) for record in concepts if record.get("scope") is not None}),
            "auxiliaryClass": sorted({str(record["class"]) for record in auxiliaries if record.get("class") is not None}),
            "accumulatorOperation": sorted({str(record["operation"]) for record in accumulators if record.get("operation") is not None}),
            "totalizes": sorted({str(record["totalizes"]) for record in concepts if record.get("totalizes") is not None}),
        },
        "duplicates": {
            "concepts": duplicate_groups(concepts, "id"),
            "auxiliaries": duplicate_groups(auxiliaries, "id"),
            "accumulators": duplicate_groups(accumulators, "id"),
        },
        "missingReferencesSample": missing[:100],
        "malformedReferencesSample": malformed_refs[:40],
        "numericConditionsSample": numeric_conditions[:80],
        "cyclesSample": cycles[:40],
        "selfEdges": [
            {"node": node, "targets": sorted(targets)}
            for node, targets in graph.items()
            if node in targets
        ],
    }

    workbook.close()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(result["counts"], ensure_ascii=False, indent=2))
    print(f"output={OUTPUT}")


if __name__ == "__main__":
    main()
